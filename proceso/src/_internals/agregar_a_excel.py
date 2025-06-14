import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


import os


def agregar_a_excel(file_path, new_data, sheet_name="Reporte"):


    if os.path.exists(file_path):
        try:
            # Cargar archivo existente
            book = load_workbook(file_path)
            if sheet_name in book.sheetnames:
                # Leer solo si la hoja ya existe
                existing_data = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
                new_data = pd.concat([existing_data, new_data], ignore_index=True)

            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                new_data.to_excel(writer, sheet_name=sheet_name, index=False)
        except InvalidFileException:
            print("⚠️ Archivo corrupto o no válido como Excel.")
    else:
        # Si no existe, simplemente escribirlo
        new_data.to_excel(file_path, sheet_name=sheet_name, index=False)