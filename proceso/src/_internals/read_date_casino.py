from openpyxl import load_workbook


def read_date_casino(input_folder, hoja="RptMovMaquinaAcumulado", celda_date="B10", celda_casino = "B7"):
    wb = load_workbook(input_folder, data_only=True)
    ws = wb[hoja]
    date = ws[celda_date].value
    date = date.split("a")[-1].strip()
    casino = ws[celda_casino].value
    casino = casino.split(":")[-1].strip()
    return date, casino